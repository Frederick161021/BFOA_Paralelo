import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.colors import ColorChoice
import openpyxl.workbook

class documentoExcel():
    def __init__(self, numeroEjecuciones, fitness, nfe, blosumscore, cpu, nucleos, memoria):
        self.numeroNucleos = len(nucleos[1])

        self.documentoExcel = openpyxl.Workbook()
        self.iniciliazarHojas()
        self.grabarDatos(numeroEjecuciones, fitness, nfe, blosumscore, cpu, nucleos, memoria)
        self.crearGrafica(self.hojaFitness, "Evaluación Fitness", "Número de Ejecucion", "Puntuación Fitness")
        self.crearGrafica(self.hojaNFE, "Evaluación NFE", "Numero de Ejecucion", "Número de Funciones Evaluadas")
        self.crearGrafica(self.hojaBlosum, "Evaluación Blosum", "Numero de Ejecucion", "Puntuaje de la Evaluación Blosum")
        self.crearGrafica(self.hojaCPU, "Uso de la CPU", "Numero de Ejecucion", "Porcentaje del Uso de la CPU")
        self.crearGrafica(self.hojaNucleos, "Uso de los Nucleos de la CPU", "Numero de Ejecucion", "Porcentaje del Uso del Nucleo")
        self.crearGrafica(self.hojaMemoria, "Uso de Memoria", "Numero de Ejecucion", "Porcentaje del Uso de Memoria")
        self.documentoExcel.save("Desempeño_Algorimto_SetC.xlsx")
        print("Se creo el documento de Excel exitosamente!")
    
    def iniciliazarHojas(self):
        self.hojaFitness = self.documentoExcel.active
        self.hojaFitness.title = "Fitness"
        self.hojaFitness.append(["Numero de Ejecución","Fitness"])

        self.hojaNFE = self.documentoExcel.create_sheet(title="NFE")
        self.hojaNFE.append(["Numero de Ejecución", "NFE"])

        self.hojaBlosum = self.documentoExcel.create_sheet(title="BlosumScore")
        self.hojaBlosum.append(["Numero de Ejecución", "BlosumScore"])

        self.hojaCPU = self.documentoExcel.create_sheet(title="CPU")
        self.hojaCPU.append(["Numero de Ejecución", "Uso del CPU"])

        self.hojaNucleos = self.documentoExcel.create_sheet(title="Nucleos")
        titulosHojaNucleo = []
        titulosHojaNucleo.append("Numero de Ejecución")
        for i in range(self.numeroNucleos):
             titulosHojaNucleo.append(f"Nucleo {i+1}")
        self.hojaNucleos.append(titulosHojaNucleo)

        self.hojaMemoria = self.documentoExcel.create_sheet(title="Memoria")
        self.hojaMemoria.append(["Numero de Ejecución", "Uso de la Memoria"])
    
    def grabarDatos(self, numeroEjecuciones, fitness, nfe, blosumScore, cpu, nucleos, memoria):
        for i in range(numeroEjecuciones):
            self.hojaFitness.append([i+1, fitness[i]])
            self.hojaNFE.append([i+1, nfe[i]])
            self.hojaBlosum.append([i+1, blosumScore[i]])
            self.hojaMemoria.append([i+1, memoria[i]])
            self.hojaCPU.append([i+1, cpu[i]])
            fila = []
            fila.append(i+1)
            for j in range (len(nucleos[i])):
                fila.append(nucleos[i][j])
            
            self.hojaNucleos.append(fila)

    def crearGrafica(self, hoja, tituloGrafica, tituloEjeX, tituloEjeY):
        grafica = LineChart()
        grafica.title = tituloGrafica
        grafica.style = 13
        grafica.x_axis.title = tituloEjeX
        grafica.y_axis.title = tituloEjeY
        
        if(tituloGrafica == "Uso de los Nucleos de la CPU"):
            datos = Reference(hoja, min_col=2, min_row=1, max_col=self.numeroNucleos+1, max_row=hoja.max_row)
        else:
            datos = Reference(hoja, min_col=2, min_row=1, max_col=2, max_row=hoja.max_row)
        
        for idx, serie in enumerate(grafica.series):
            serie.graphicalProperties.line.solidFill = self.obtenerColorPorIndice(idx)

        grafica.add_data(datos, titles_from_data=True)
        hoja.add_chart(grafica, "F5")

    

# arreglo = [1,2,3,4,5]
# arreglo2 = [10,8,9,7,10] 

# documentoExcel(5, arreglo, arreglo, arreglo2, arreglo2, arreglo2)